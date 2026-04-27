"""Sync drones from products.xlsx into the database."""
import os
from pathlib import Path

EXCEL_PATH = Path(__file__).resolve().parent.parent / 'products.xlsx'


def sync_drones_from_excel():
    """Read products.xlsx and upsert Drone records."""
    try:
        import openpyxl
    except ImportError:
        return

    if not EXCEL_PATH.exists():
        return

    from .models import Drone

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    def col(row, name):
        try:
            idx = headers.index(name)
            return row[idx].value
        except (ValueError, IndexError):
            return None

    for row in ws.iter_rows(min_row=2):
        name = col(row, 'name')
        if not name:
            continue
        price = col(row, 'price') or 0
        stock = col(row, 'stock') or 0
        description = col(row, 'description') or ''
        category = col(row, 'category') or ''
        weight_kg = col(row, 'weight_kg') or 0
        flight_time_min = col(row, 'flight_time_min') or 0
        range_km = col(row, 'range_km') or 0

        Drone.objects.update_or_create(
            name=name,
            defaults={
                'price': price,
                'stock': stock,
                'description': description,
                'category': category,
                'weight_kg': weight_kg,
                'flight_time_min': flight_time_min,
                'range_km': range_km,
                'available': stock > 0,
            }
        )
